from ..domain import LancamentoCreate, LancamentoOut
from ..auth.contexto import cod_pessoa_atual
from ..repositories import LancamentoRepository
import csv
import openpyxl
from io import BytesIO, StringIO
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..models import (
    db,
    Lancamento,
    Qualificador,
    TipoLancamento,
    OrigemLancamento,
    ContaBancaria,
    Conferencia,
)


def list_lancamentos(
    start_date: date | None = None,
    end_date: date | None = None,
    tipo: int | None = None,
    qualificador_folha: int | None = None,
    seq_conta: int | None = None,
    cod_origem: int | None = None,
    page: int = 1,
    per_page: int = 50,
    sort_by: str = 'dat_lancamento',
    sort_order: str = 'desc',
    repo: LancamentoRepository | None = None
) -> tuple[list, int]:
    repo = repo or LancamentoRepository()
    return repo.list(
        start_date=start_date,
        end_date=end_date,
        tipo=tipo,
        qualificador_folha=qualificador_folha,
        seq_conta=seq_conta,
        cod_origem=cod_origem,
        page=page,
        per_page=per_page,
        sort_by=sort_by,
        sort_order=sort_order
    )


def list_tipos_lancamento():
    from ..repositories import TipoLancamentoRepository
    repo = TipoLancamentoRepository()
    return repo.list_all()


def list_origens_lancamento():
    from ..repositories import OrigemLancamentoRepository
    repo = OrigemLancamentoRepository()
    return repo.list_all()


def list_contas_bancarias():
    from ..repositories import ContaBancariaRepository
    repo = ContaBancariaRepository()
    return repo.list_active()


def list_conferencias():
    from ..repositories import ConferenciaRepository
    repo = ConferenciaRepository()
    return repo.list_all()


def _validar_dados_lancamento(data: LancamentoCreate) -> None:
    """Regras de negócio de lançamento (spec cadastros-nucleo R2)."""
    from decimal import Decimal

    from .validacao import RegraNegocioError

    # F6.1b: o sinal do fluxo de caixa vive no tipo ('C'/'D'), então o valor
    # é sempre positivo. Antes daqui só se exigia != 0, e receita negativa /
    # despesa positiva entravam sem barreira.
    if Decimal(str(data.val_lancamento)) <= 0:
        raise RegraNegocioError(
            "O valor do lançamento deve ser positivo — o tipo (Entrada/Saída) "
            "é que define o sinal no fluxo de caixa"
        )

    qualificador = Qualificador.query.get(data.seq_qualificador)
    if qualificador is None or qualificador.ind_status != 'A':
        raise RegraNegocioError("Qualificador inexistente ou inativo")
    if not qualificador.is_folha():
        raise RegraNegocioError("Lançamentos só podem ser feitos em qualificadores folha")

    if TipoLancamento.query.get(data.cod_tipo_lancamento) is None:
        raise RegraNegocioError("Tipo de lançamento inexistente")
    if OrigemLancamento.query.get(data.cod_origem_lancamento) is None:
        raise RegraNegocioError("Origem de lançamento inexistente")
    if data.seq_conta is not None and ContaBancaria.query.get(data.seq_conta) is None:
        raise RegraNegocioError("Conta bancária inexistente")


def _exigir_origem_manual(lanc) -> None:
    from .validacao import RegraNegocioError

    origem = OrigemLancamento.query.get(lanc.cod_origem_lancamento)
    if origem is not None and origem.dsc_origem_lancamento != 'Manual':
        raise RegraNegocioError(
            f"Lançamentos de origem {origem.dsc_origem_lancamento} "
            "não podem ser alterados ou excluídos"
        )


def create_lancamento(data: LancamentoCreate, repo: LancamentoRepository | None = None) -> LancamentoOut:
    repo = repo or LancamentoRepository()
    _validar_dados_lancamento(data)
    lanc = repo.create(data)
    return LancamentoOut(
        seq_lancamento=lanc.seq_lancamento,
        dat_lancamento=lanc.dat_lancamento,
        seq_qualificador=lanc.seq_qualificador,
        val_lancamento=lanc.val_lancamento,
        cod_tipo_lancamento=lanc.cod_tipo_lancamento,
        cod_origem_lancamento=lanc.cod_origem_lancamento,
        dsc_lancamento=None,
    seq_conta=lanc.seq_conta,
    )


def update_lancamento(ident: int, data: LancamentoCreate, repo: LancamentoRepository | None = None) -> LancamentoOut:
    from .validacao import RegraNegocioError

    repo = repo or LancamentoRepository()
    atual = Lancamento.query.get(ident)
    if atual is None:
        raise RegraNegocioError("Lançamento inexistente")
    _exigir_origem_manual(atual)
    if data.dat_lancamento != atual.dat_lancamento:
        raise RegraNegocioError("A data do lançamento não pode ser alterada")
    _validar_dados_lancamento(data)
    lanc = repo.update(ident, data)
    return LancamentoOut(
        seq_lancamento=lanc.seq_lancamento,
        dat_lancamento=lanc.dat_lancamento,
        seq_qualificador=lanc.seq_qualificador,
        val_lancamento=lanc.val_lancamento,
        cod_tipo_lancamento=lanc.cod_tipo_lancamento,
        cod_origem_lancamento=lanc.cod_origem_lancamento,
        dsc_lancamento=None,
    seq_conta=lanc.seq_conta,
    )


def delete_lancamento(ident: int, repo: LancamentoRepository | None = None):
    from .validacao import RegraNegocioError

    repo = repo or LancamentoRepository()
    atual = Lancamento.query.get(ident)
    if atual is None:
        raise RegraNegocioError("Lançamento inexistente")
    _exigir_origem_manual(atual)
    repo.soft_delete(ident)


def import_lancamentos_service(
    content: bytes, filename: str, session: Session | None = None
) -> dict:
    session = session or db.session
    rows: list[dict] = []
    
    if filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(StringIO(text))
        for row in reader:
            rows.append({k.strip(): v for k, v in row.items()})
    elif filename.lower().endswith(('.xlsx', '.xls')):
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c).strip() if c else '' for c in next(ws.iter_rows(values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            rows.append(data)
    else:
        return {"sucesso": 0, "erros": ["Formato de arquivo não suportado"]}

    # Pre-fetch data to avoid N+1 queries
    all_qualificadores = session.query(Qualificador).all()
    qualificadores_map = {q.dsc_qualificador.lower(): q for q in all_qualificadores}

    all_tipos = session.query(TipoLancamento).all()
    # F6.1b: aceita a descrição ("Entrada"/"Saída") e o código ('C'/'D').
    # O código numérico antigo (1/2) não resolve mais nada e é recusado.
    tipos_map = {t.dsc_tipo_lancamento.lower(): t.cod_tipo_lancamento for t in all_tipos}
    tipos_map.update({t.cod_tipo_lancamento.lower(): t.cod_tipo_lancamento for t in all_tipos})

    all_origens = session.query(OrigemLancamento).all()
    origens_map = {o.dsc_origem_lancamento.lower(): o.cod_origem_lancamento for o in all_origens}
    
    # Use "Importado" origin for imported entries
    origem_importado = session.query(OrigemLancamento).filter_by(dsc_origem_lancamento='Importado').first()
    origem_importado_cod = origem_importado.cod_origem_lancamento if origem_importado else None

    all_contas = session.query(ContaBancaria).all()
    contas_map = {(c.cod_banco, c.num_agencia, c.num_conta): c for c in all_contas}

    count = 0
    errors = []

    def get_or_create_conta(banco, agencia, conta):
        if not (banco and agencia and conta):
            return None
        banco = str(banco).strip()
        agencia = str(agencia).strip()
        conta = str(conta).strip()
        
        key = (banco, agencia, conta)
        if key in contas_map:
            return contas_map[key]
            
        c = ContaBancaria(
            cod_banco=banco,
            num_agencia=agencia,
            num_conta=conta,
            dsc_conta=f"{banco}-{agencia}/{conta}",
            cod_pessoa_inclusao=cod_pessoa_atual(),
        )
        session.add(c)
        session.flush()
        contas_map[key] = c
        return c


    try:
        for i, item in enumerate(rows, start=2):
            dat = item.get('Data') or item.get('dat_lancamento')
            desc = item.get('Qualificador') or item.get('Descrição') or item.get('descricao')
            valor = item.get('Valor (R$)') or item.get('val_lancamento')
            tipo_raw = item.get('Tipo') or item.get('cod_tipo_lancamento')

            if not (dat and desc and valor and tipo_raw):
                errors.append(f"Linha {i}: Dados incompletos (Data, Qualificador, Valor ou Tipo faltando)")
                continue

            if isinstance(dat, datetime):
                dat = dat.date()
            elif isinstance(dat, str):
                try:
                    dat = date.fromisoformat(dat)
                except ValueError:
                    errors.append(f"Linha {i}: Data inválida '{dat}'")
                    continue

            qual = qualificadores_map.get(str(desc).lower())
            if not qual:
                errors.append(f"Linha {i}: Qualificador não encontrado para '{desc}'")
                continue

            tipo = tipos_map.get(str(tipo_raw).strip().lower())
            if not tipo:
                errors.append(
                    f"Linha {i}: Tipo inválido '{tipo_raw}' — use "
                    f"Entrada/Saída ou C/D"
                )
                continue

            try:
                valor_dec = Decimal(str(valor))
            except (InvalidOperation, ValueError):
                errors.append(f"Linha {i}: Valor inválido '{valor}'")
                continue
            if valor_dec <= 0:
                errors.append(
                    f"Linha {i}: O valor deve ser positivo — o tipo "
                    f"(Entrada/Saída) é que define o sinal"
                )
                continue

            # Use Importado origin for all imported entries
            if not origem_importado_cod:
                errors.append(f"Linha {i}: Origem 'Importado' não encontrada no sistema")
                continue

            # Detect optional bank fields
            banco = item.get('Banco') or item.get('banco') or item.get('BANCO')
            agencia = item.get('Agencia') or item.get('agencia') or item.get('AGENCIA')
            conta = item.get('Conta') or item.get('conta') or item.get('CONTA')
            conta_obj = get_or_create_conta(banco, agencia, conta)

            lanc = Lancamento(
                dat_lancamento=dat,
                seq_qualificador=qual.seq_qualificador,
                val_lancamento=valor_dec,
                cod_tipo_lancamento=tipo,
                cod_origem_lancamento=origem_importado_cod,
                cod_pessoa_inclusao=cod_pessoa_atual(),
                seq_conta=(conta_obj.seq_conta if conta_obj else None),
            )
            session.add(lanc)
            count += 1
        
        session.commit()
        return {"sucesso": count, "erros": errors}
    except Exception as e:
        session.rollback()
        errors.append(f"Erro fatal durante importação: {str(e)}")
        return {"sucesso": 0, "erros": errors}
