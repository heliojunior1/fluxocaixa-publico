"""Fachada da tela de saldos sobre o modelo por fundo (spec R19/R20).

Após o corte do legado (F2.4), a tela grava/edita/inativa no modelo novo
(`gravar_saldo`/`inativar_saldo`) e importa CSV via `importar_lote` no fundo
GERAL (tipo IMPORTADO). A leitura para a listagem usa as views:
agregado por conta (default) ou por fundo.
"""
import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

import openpyxl

from ..models import ContaBancaria, Fundo, SistemaOrigem
from ..models.base import db
from ..services.fundo_service import garantir_fundo_geral
from ..services.importacao_lote_service import LinhaLote, importar_lote
from ..services.saldo_fundo_service import gravar_saldo, inativar_saldo

__all__ = [
    "listar_saldos_tela", "criar_saldo_tela", "editar_saldo_tela",
    "inativar_saldo_tela", "inativar_saldo", "gravar_saldo",
]


def listar_saldos_tela(visao: str = "agregado", seq_conta=None,
                       data_inicio=None, data_fim=None, seq_fundo=None) -> list[dict]:
    """Lista saldos para a tela: 'agregado' (soma por conta/dia) ou 'fundo'."""
    if visao == "fundo":
        return _listar_por_fundo(seq_conta, data_inicio, data_fim, seq_fundo)
    return _listar_agregado(seq_conta, data_inicio, data_fim)


def _filtro_datas(sql, params, data_inicio, data_fim):
    if data_inicio:
        sql += " AND dat_saldo >= :ini"; params["ini"] = data_inicio
    if data_fim:
        sql += " AND dat_saldo <= :fim"; params["fim"] = data_fim
    return sql, params


def _listar_agregado(seq_conta, data_inicio, data_fim) -> list[dict]:
    from sqlalchemy import text

    sql = ("SELECT seq_conta, dat_saldo, val_saldo, val_aplicacoes, val_resgates, "
           "dsc_origem_consolidada FROM vw_flc_saldo_conta_agregado WHERE 1=1")
    params = {}
    if seq_conta:
        sql += " AND seq_conta = :c"; params["c"] = seq_conta
    sql, params = _filtro_datas(sql, params, data_inicio, data_fim)
    sql += " ORDER BY dat_saldo DESC, seq_conta"
    contas = {c.seq_conta: c for c in ContaBancaria.query.all()}
    linhas = []
    for r in db.session.execute(text(sql), params).mappings():
        dat = r["dat_saldo"]
        linhas.append({
            "seq_conta": r["seq_conta"],
            "conta": contas.get(r["seq_conta"]),
            "dat_saldo": date.fromisoformat(dat) if isinstance(dat, str) else dat,
            "val_saldo": Decimal(str(r["val_saldo"] or 0)),
            "origem": r["dsc_origem_consolidada"],
        })
    return linhas


def _listar_por_fundo(seq_conta, data_inicio, data_fim, seq_fundo) -> list[dict]:
    from ..models import SaldoContaFundo

    q = SaldoContaFundo.query.filter_by(ind_status='A')
    if seq_conta:
        q = q.filter_by(seq_conta=seq_conta)
    if seq_fundo:
        q = q.filter_by(seq_fundo=seq_fundo)
    if data_inicio:
        q = q.filter(SaldoContaFundo.dat_saldo >= data_inicio)
    if data_fim:
        q = q.filter(SaldoContaFundo.dat_saldo <= data_fim)
    contas = {c.seq_conta: c for c in ContaBancaria.query.all()}
    fundos = {f.seq_fundo: f for f in Fundo.query.all()}
    linhas = []
    for s in q.order_by(SaldoContaFundo.dat_saldo.desc()).all():
        linhas.append({
            "seq_conta": s.seq_conta, "conta": contas.get(s.seq_conta),
            "seq_fundo": s.seq_fundo, "fundo": fundos.get(s.seq_fundo),
            "dat_saldo": s.dat_saldo, "val_saldo": Decimal(str(s.val_saldo)),
            "val_aplicacoes": Decimal(str(s.val_aplicacoes)),
            "val_resgates": Decimal(str(s.val_resgates)),
        })
    return linhas


def criar_saldo_tela(seq_conta, seq_fundo, dat_saldo, val_saldo,
                     val_aplicacoes="0", val_resgates="0", sigla_sistema=None):
    tipo = 'AUTOMATIZADO' if sigla_sistema else 'MANUAL'
    return gravar_saldo(
        seq_conta=seq_conta, seq_fundo=seq_fundo, dat_saldo=dat_saldo,
        val_saldo=Decimal(str(val_saldo)),
        val_aplicacoes=Decimal(str(val_aplicacoes)),
        val_resgates=Decimal(str(val_resgates)),
        sigla_tipo_origem=tipo, sigla_sistema_origem=sigla_sistema,
    )


def editar_saldo_tela(seq_conta, seq_fundo, dat_saldo, val_saldo,
                      val_aplicacoes="0", val_resgates="0"):
    """Edição = regravação (inativa a anterior + insere) — chaves imutáveis."""
    return gravar_saldo(
        seq_conta=seq_conta, seq_fundo=seq_fundo, dat_saldo=dat_saldo,
        val_saldo=Decimal(str(val_saldo)),
        val_aplicacoes=Decimal(str(val_aplicacoes)),
        val_resgates=Decimal(str(val_resgates)),
    )


def inativar_saldo_tela(seq_conta, seq_fundo, dat_saldo):
    return inativar_saldo(seq_conta, seq_fundo, dat_saldo)


def _parse_arquivo(content: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith('.csv'):
        # aceita ; ou , como separador
        texto = content.decode('utf-8-sig')
        sep = ';' if ';' in texto.splitlines()[0] else ','
        return [dict(r) for r in csv.DictReader(StringIO(texto), delimiter=sep)]
    if filename.lower().endswith(('.xlsx', '.xls')):
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c).strip() if c else '' for c in next(ws.iter_rows(values_only=True))]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            out.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        return out
    return []


def _parse_data(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in (None, '%d/%m/%Y'):
            try:
                return date.fromisoformat(v) if fmt is None else datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None
